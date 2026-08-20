import io

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from businesses.models import Business
from payroll.models import Employee, Payment


class PaymentExportAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.business = Business.objects.create(business_name="카페비서")
        self.url = f"/api/businesses/{self.business.id}/payroll/payments/export/"
        employee = Employee.objects.create(
            business=self.business, name="장예은", employment_type="FULL_TIME", hourly_wage=10320
        )
        Payment.objects.create(
            employee=employee, year=2026, month=8,
            work_hours=141, gross_pay=1_455_120, withholding_tax=8_734,
        )
        part_timer = Employee.objects.create(
            business=self.business,
            name="박서연",
            employment_type="PART_TIME",
            hourly_wage=10200,
            is_long_term_contract=True,
        )
        Payment.objects.create(
            employee=part_timer,
            year=2026,
            month=8,
            work_hours=80,
            gross_pay=816_000,
            withholding_tax=0,
        )
        uninsured_part_timer = Employee.objects.create(
            business=self.business,
            name="이단기",
            employment_type="PART_TIME",
            hourly_wage=10200,
            is_long_term_contract=False,
        )
        Payment.objects.create(
            employee=uninsured_part_timer,
            year=2026,
            month=8,
            work_hours=80,
            gross_pay=816_000,
            withholding_tax=0,
        )

    def test_export_pdf_returns_pdf_file(self):
        response = self.client.post(self.url, {"year": 2026, "month": 8, "format": "pdf"}, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertGreater(len(response.content), 0)
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_export_xlsx_returns_xlsx_file(self):
        response = self.client.post(self.url, {"year": 2026, "month": 8, "format": "xlsx"}, format="json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertGreater(len(response.content), 0)
        self.assertTrue(response.content.startswith(b"PK"))

    def test_export_xlsx_includes_part_time_employment_insurance(self):
        response = self.client.post(
            self.url,
            {"year": 2026, "month": 8, "format": "xlsx"},
            format="json",
        )

        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        rows = {
            sheet.cell(row=row, column=headers["직원명"]).value: row
            for row in range(2, sheet.max_row + 1)
        }
        part_time_row = rows["박서연"]

        self.assertEqual(
            sheet.cell(row=part_time_row, column=headers["고용보험"]).value,
            round(816_000 * 0.009),
        )
        uninsured_row = rows["이단기"]
        self.assertEqual(
            sheet.cell(row=uninsured_row, column=headers["고용보험"]).value,
            0,
        )
        self.assertGreater(
            sheet.cell(row=part_time_row, column=headers["공제액 합계"]).value,
            sheet.cell(row=uninsured_row, column=headers["공제액 합계"]).value,
        )

    def test_export_invalid_format_returns_400(self):
        response = self.client.post(self.url, {"year": 2026, "month": 8, "format": "hwp"}, format="json")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["code"], "INVALID_EXPORT_FORMAT")

    def test_export_no_data_for_period_returns_404(self):
        response = self.client.post(self.url, {"year": 2020, "month": 1, "format": "pdf"}, format="json")
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data["code"], "PAYROLL_DATA_NOT_FOUND")
