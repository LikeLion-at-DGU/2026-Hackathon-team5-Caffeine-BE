"""임금명세서/지급명세서 엑셀(XLSX) 일괄 내보내기.

PDF(직원별 개별 서식)와 달리, 엑셀은 직원별로 한 행씩 나열하는 표 형태로 만든다.
사장님이 엑셀에서 정렬/필터링해서 보기 편하도록 하는 목적.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from payroll.services.payment_service import get_payslip_data

_HEADERS = [
    ("직원명", "employee_name"),
    ("고용형태", "employment_type"),
    ("근무시간", "work_hours"),
    ("시급", "hourly_wage"),
    ("세전급여", "gross_pay"),
    ("소득세", "income_tax"),
    ("지방소득세", "local_income_tax"),
    ("국민연금", "national_pension"),
    ("건강보험", "health_insurance"),
    ("장기요양보험", "long_term_care"),
    ("고용보험", "employment_insurance"),
    ("공제액 합계", "deductions_total"),
    ("실수령액", "net_pay"),
]

_EMPLOYMENT_TYPE_LABELS = {
    "FULL_TIME": "4대보험 정직원",
    "PART_TIME": "단시간 근로자",
    "FREELANCER": "3.3% 프리랜서",
}


def generate_payslip_xlsx(payments: list, year: int, month: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}년 {month}월 급여명세서"

    header_font = Font(bold=True)
    for col_idx, (header_label, _) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, payment in enumerate(payments, start=2):
        data = get_payslip_data(payment)
        data["employment_type"] = _EMPLOYMENT_TYPE_LABELS.get(data["employment_type"], data["employment_type"])
        for col_idx, (_, field_key) in enumerate(_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=data[field_key])

    for col_idx in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()